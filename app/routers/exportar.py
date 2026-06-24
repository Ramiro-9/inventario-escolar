from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from app.database import get_db
from app import models
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

router = APIRouter(prefix="/exportar", tags=["Exportar"])

def _get_datos(db: Session):
    stock    = db.get(models.Stock, 1)
    cursos   = db.query(models.Curso).order_by(models.Curso.nombre).all()
    resumen  = db.execute(text("SELECT * FROM resumen_inventario")).mappings().first()
    req_ubi  = db.execute(text("SELECT * FROM requerimientos_por_ubicacion ORDER BY ubicacion")).mappings().all()
    return stock, cursos, resumen, req_ubi

# ── Excel ──────────────────────────────────────────────────────────────────
@router.get("/excel")
def exportar_excel(db: Session = Depends(get_db)):
    stock, cursos, resumen, req_ubi = _get_datos(db)

    wb = openpyxl.Workbook()

    # Estilos
    hdr  = Font(bold=True, color="FFFFFF")
    fill_dark  = PatternFill("solid", fgColor="1A1A2E")
    fill_mid   = PatternFill("solid", fgColor="3730A3")
    fill_green = PatternFill("solid", fgColor="166534")
    fill_red   = PatternFill("solid", fgColor="991B1B")
    fill_gray  = PatternFill("solid", fgColor="E5E7EB")
    center     = Alignment(horizontal="center", vertical="center")
    thin       = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def hdr_row(ws, row, cols, fill):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = fill
            cell.alignment = center
            cell.border    = thin

    def data_row(ws, row, vals, alt=False):
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = center
            cell.border    = thin
            if alt: cell.fill = fill_gray

    # ── Hoja 1: Resumen global ──
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    ws1.merge_cells("A1:B1")
    t = ws1["A1"]; t.value = "INVENTARIO ESCOLAR — RESUMEN GLOBAL"
    t.font = Font(bold=True, size=13, color="FFFFFF"); t.fill = fill_dark
    t.alignment = center

    ws1.merge_cells("A2:B2")
    f = ws1["A2"]; f.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    f.alignment = center; f.font = Font(italic=True, color="6B7280")

    datos = [
        ("Bancos disponibles",  resumen["bancos_total"]),
        ("Bancos requeridos",   resumen["bancos_requeridos"]),
        ("Balance bancos",      resumen["bancos_sobrantes"]),
        ("Sillas disponibles",  resumen["sillas_total"]),
        ("Sillas requeridas",   resumen["sillas_requeridas"]),
        ("Balance sillas",      resumen["sillas_sobrantes"]),
    ]
    for i, (label, val) in enumerate(datos, 4):
        ws1.cell(row=i, column=1, value=label).border = thin
        cell = ws1.cell(row=i, column=2, value=val)
        cell.alignment = center; cell.border = thin
        if "Balance" in label:
            cell.fill = fill_green if val >= 0 else fill_red
            cell.font = Font(bold=True, color="FFFFFF")

    # ── Hoja 2: Por ubicación ──
    ws2 = wb.create_sheet("Por Aula")
    ws2.column_dimensions["A"].width = 28
    for col in ["B","C","D"]: ws2.column_dimensions[col].width = 18

    hdr_row(ws2, 1, ["Ubicación", "Cursos", "Bancos req. (MAX)", "Sillas req. (MAX)"], fill_mid)
    for i, r in enumerate(req_ubi, 2):
        data_row(ws2, i,
            [r["ubicacion"], r["cantidad_cursos"] or 0,
             r["bancos_requeridos"] or 0, r["sillas_requeridas"] or 0],
            alt=(i % 2 == 0)
        )

    # ── Hoja 3: Cursos ──
    ws3 = wb.create_sheet("Cursos")
    for col, w in zip(["A","B","C","D","E"], [18,14,28,18,18]):
        ws3.column_dimensions[col].width = w

    hdr_row(ws3, 1, ["Curso", "Turno", "Ubicación", "Bancos req.", "Sillas req."], fill_mid)
    for i, c in enumerate(cursos, 2):
        data_row(ws3, i,
            [c.nombre, c.turno.value if c.turno else "—",
             c.ubicacion.nombre, c.bancos_requeridos, c.sillas_requeridas],
            alt=(i % 2 == 0)
        )

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# ── PDF ────────────────────────────────────────────────────────────────────
@router.get("/pdf")
def exportar_pdf(db: Session = Depends(get_db)):
    stock, cursos, resumen, req_ubi = _get_datos(db)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    COLOR_DARK  = colors.HexColor("#1A1A2E")
    COLOR_ACC   = colors.HexColor("#6366F1")
    COLOR_GREEN = colors.HexColor("#166534")
    COLOR_RED   = colors.HexColor("#991B1B")
    COLOR_GRAY  = colors.HexColor("#F3F4F6")
    COLOR_WHITE = colors.white

    def section(title):
        story.append(Spacer(1, 0.4*cm))
        p = Paragraph(f"<b>{title}</b>",
            styles["Heading2"] if False else getSampleStyleSheet()["Normal"])
        p.style.fontSize     = 12
        p.style.textColor    = COLOR_ACC
        p.style.spaceAfter   = 6
        story.append(p)

    def make_table(data, col_widths, hdr_color=COLOR_DARK):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), hdr_color),
            ("TEXTCOLOR",  (0,0), (-1,0), COLOR_WHITE),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [COLOR_WHITE, COLOR_GRAY]),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ]))
        return t

    # Título
    title = Paragraph(
        "<b>INVENTARIO ESCOLAR — EPET N°1 CAUCETE</b>",
        getSampleStyleSheet()["Title"]
    )
    story.append(title)
    story.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        getSampleStyleSheet()["Normal"]
    ))
    story.append(Spacer(1, 0.5*cm))

    # Resumen global
    section("Resumen Global")
    bsob = resumen["bancos_sobrantes"]
    ssob = resumen["sillas_sobrantes"]
    data_res = [
        ["", "Bancos", "Sillas"],
        ["Disponibles", resumen["bancos_total"],      resumen["sillas_total"]],
        ["Requeridos",  resumen["bancos_requeridos"],  resumen["sillas_requeridas"]],
        ["Balance",     f"+{bsob}" if bsob>=0 else str(bsob),
                        f"+{ssob}" if ssob>=0 else str(ssob)],
    ]
    t_res = make_table(data_res, [6*cm, 4*cm, 4*cm])
    # Color balance
    t_res.setStyle(TableStyle([
        ("BACKGROUND", (1,3), (1,3), COLOR_GREEN if bsob >= 0 else COLOR_RED),
        ("BACKGROUND", (2,3), (2,3), COLOR_GREEN if ssob >= 0 else COLOR_RED),
        ("TEXTCOLOR",  (1,3), (2,3), COLOR_WHITE),
        ("FONTNAME",   (1,3), (2,3), "Helvetica-Bold"),
    ]))
    story.append(t_res)

    # Por aula
    section("Requerimientos por Aula")
    data_ubi = [["Ubicación", "Cursos", "Bancos req.", "Sillas req."]]
    for r in req_ubi:
        data_ubi.append([
            r["ubicacion"], r["cantidad_cursos"] or 0,
            r["bancos_requeridos"] or 0, r["sillas_requeridas"] or 0
        ])
    story.append(make_table(data_ubi, [7*cm, 3*cm, 3.5*cm, 3.5*cm]))

    # Cursos
    section("Listado de Cursos")
    data_cur = [["Curso", "Turno", "Ubicación", "Bancos", "Sillas"]]
    for c in cursos:
        data_cur.append([
            c.nombre, c.turno.value if c.turno else "—",
            c.ubicacion.nombre, c.bancos_requeridos, c.sillas_requeridas
        ])
    story.append(make_table(data_cur, [3*cm, 2.5*cm, 7*cm, 2.5*cm, 2.5*cm]))

    doc.build(story)
    buf.seek(0)
    fname = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )